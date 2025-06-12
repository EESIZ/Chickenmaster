#!/usr/bin/env python3
"""
엑셀 파일에 상수 관리 시트들을 추가하는 스크립트

기존 game_initial_values_with_formulas.xlsx 파일에 
상수 관리를 위한 새로운 시트들을 추가합니다.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_constants_sheets_data() -> Dict[str, pd.DataFrame]:
    """상수 관리 시트들의 데이터를 생성합니다."""
    
    # Core_Constants: 핵심 게임 상수
    core_constants_data = [
        # 확률 관련
        {"Key": "PROBABILITY_HIGH_THRESHOLD", "Value": 0.7, "Type": "float", 
         "Description": "높은 확률 임계값 (70%)", "Category": "probability"},
        {"Key": "PROBABILITY_MEDIUM_THRESHOLD", "Value": 0.5, "Type": "float", 
         "Description": "중간 확률 임계값 (50%)", "Category": "probability"},
        {"Key": "PROBABILITY_LOW_THRESHOLD", "Value": 0.3, "Type": "float", 
         "Description": "낮은 확률 임계값 (30%)", "Category": "probability"},
        
        # 게임 진행
        {"Key": "TOTAL_GAME_DAYS", "Value": 730, "Type": "int", 
         "Description": "총 게임 일수 (2년)", "Category": "game_flow"},
        {"Key": "MAX_ACTIONS_PER_DAY", "Value": 3, "Type": "int", 
         "Description": "하루 최대 행동 횟수", "Category": "game_flow"},
        {"Key": "EARLY_GAME_THRESHOLD", "Value": 180, "Type": "int", 
         "Description": "초기 게임 단계 임계값 (약 6개월)", "Category": "game_flow"},
        {"Key": "MID_GAME_THRESHOLD", "Value": 545, "Type": "int", 
         "Description": "중반 게임 단계 임계값 (약 1년 6개월)", "Category": "game_flow"},
        
        # 경제 관련
        {"Key": "DEFAULT_STARTING_MONEY", "Value": 10000.0, "Type": "float", 
         "Description": "게임 시작 시 기본 자금", "Category": "economy"},
        {"Key": "CHICKEN_INGREDIENT_COST", "Value": 5681.0, "Type": "float", 
         "Description": "치킨 1마리 재료비 (2025년 기준)", "Category": "economy"},
        {"Key": "STARTUP_COST_2025", "Value": 68150000.0, "Type": "float", 
         "Description": "치킨집 창업 비용 (2025년 기준)", "Category": "economy"},
        
        # 임계값
        {"Key": "MONEY_LOW_THRESHOLD", "Value": 3000.0, "Type": "float", 
         "Description": "자금 부족 경고 기준", "Category": "thresholds"},
        {"Key": "MONEY_HIGH_THRESHOLD", "Value": 15000.0, "Type": "float", 
         "Description": "자금 풍부 기준", "Category": "thresholds"},
        {"Key": "REPUTATION_LOW_THRESHOLD", "Value": 30.0, "Type": "float", 
         "Description": "평판 위험 기준", "Category": "thresholds"},
        {"Key": "REPUTATION_HIGH_THRESHOLD", "Value": 70.0, "Type": "float", 
         "Description": "평판 우수 기준", "Category": "thresholds"},
        {"Key": "HAPPINESS_LOW_THRESHOLD", "Value": 30.0, "Type": "float", 
         "Description": "행복 위험 기준", "Category": "thresholds"},
        {"Key": "HAPPINESS_HIGH_THRESHOLD", "Value": 70.0, "Type": "float", 
         "Description": "행복 우수 기준", "Category": "thresholds"},
    ]
    
    # Magic_Numbers: 현재 하드코딩된 매직넘버들
    magic_numbers_data = [
        {"Key": "FLOAT_EPSILON", "Value": 0.001, "Type": "float", 
         "Description": "부동소수점 비교 오차 허용 범위", "Category": "math"},
        {"Key": "MAX_CASCADE_DEPTH", "Value": 5, "Type": "int", 
         "Description": "최대 연쇄 효과 깊이", "Category": "events"},
        {"Key": "MAX_CASCADE_NODES", "Value": 100, "Type": "int", 
         "Description": "최대 연쇄 노드 수", "Category": "events"},
        {"Key": "EVENT_COOLDOWN_DAYS", "Value": 7, "Type": "int", 
         "Description": "이벤트 쿨다운 일수", "Category": "events"},
        {"Key": "MAX_EVENTS_PER_DAY", "Value": 3, "Type": "int", 
         "Description": "하루 최대 이벤트 수", "Category": "events"},
        {"Key": "REPUTATION_BASELINE", "Value": 50, "Type": "int", 
         "Description": "평판 기준점", "Category": "metrics"},
        {"Key": "HAPPINESS_SUFFERING_SUM", "Value": 100.0, "Type": "float", 
         "Description": "행복도 + 고통도 = 100 (시소 불변식)", "Category": "metrics"},
        {"Key": "MAX_INVENTORY_SIZE", "Value": 10, "Type": "int", 
         "Description": "최대 인벤토리 크기", "Category": "inventory"},
        {"Key": "MAX_ITEM_QUANTITY", "Value": 99, "Type": "int", 
         "Description": "최대 아이템 수량", "Category": "inventory"},
    ]
    
    # Test_Constants: 테스트용 상수들
    test_constants_data = [
        {"Key": "TEST_MONEY", "Value": 20000.0, "Type": "float", 
         "Description": "테스트용 자금", "Category": "test_data"},
        {"Key": "TEST_REPUTATION", "Value": 75.0, "Type": "float", 
         "Description": "테스트용 평판", "Category": "test_data"},
        {"Key": "TEST_HAPPINESS", "Value": 80.0, "Type": "float", 
         "Description": "테스트용 행복도", "Category": "test_data"},
        {"Key": "TEST_MIN_CASCADE_EVENTS", "Value": 3, "Type": "int", 
         "Description": "최소 연쇄 효과 메시지 수", "Category": "test_validation"},
        {"Key": "TEST_EXPECTED_EVENTS", "Value": 2, "Type": "int", 
         "Description": "예상 이벤트 수", "Category": "test_validation"},
        {"Key": "TEST_METRICS_HISTORY_LENGTH", "Value": 5, "Type": "int", 
         "Description": "메트릭 히스토리 길이", "Category": "test_validation"},
        {"Key": "MAX_RETRY_ATTEMPTS", "Value": 3, "Type": "int", 
         "Description": "최대 재시도 횟수", "Category": "test_config"},
        {"Key": "TIMEOUT_SECONDS", "Value": 1.0, "Type": "float", 
         "Description": "재시도 간 대기 시간(초)", "Category": "test_config"},
    ]
    
    # UI_Constants: 화면 표시 관련 상수
    ui_constants_data = [
        {"Key": "MIN_STORY_LENGTH", "Value": 100, "Type": "int", 
         "Description": "최소 스토리 길이", "Category": "ui_display"},
        {"Key": "MAX_STORY_LENGTH", "Value": 1000, "Type": "int", 
         "Description": "최대 스토리 길이", "Category": "ui_display"},
        {"Key": "RECENT_HISTORY_WINDOW", "Value": 3, "Type": "int", 
         "Description": "최근 히스토리 분석 윈도우 크기", "Category": "ui_display"},
        {"Key": "MIN_METRICS_HISTORY_FOR_TREND", "Value": 2, "Type": "int", 
         "Description": "추세 분석을 위한 최소 히스토리 개수", "Category": "ui_analysis"},
        {"Key": "MINIMUM_TREND_POINTS", "Value": 2, "Type": "int", 
         "Description": "트렌드 분석에 필요한 최소 데이터 포인트", "Category": "ui_analysis"},
    ]
    
    # Performance_Constants: 성능 관련 설정
    performance_constants_data = [
        {"Key": "CACHE_EXPIRY_SECONDS", "Value": 300, "Type": "int", 
         "Description": "캐시 만료 시간 (5분)", "Category": "caching"},
        {"Key": "MAX_CONCURRENT_EVENTS", "Value": 10, "Type": "int", 
         "Description": "최대 동시 처리 이벤트 수", "Category": "performance"},
        {"Key": "BATCH_SIZE", "Value": 50, "Type": "int", 
         "Description": "배치 처리 크기", "Category": "performance"},
        {"Key": "MEMORY_LIMIT_MB", "Value": 512, "Type": "int", 
         "Description": "메모리 사용 제한 (MB)", "Category": "performance"},
        {"Key": "LOG_ROTATION_SIZE_MB", "Value": 10, "Type": "int", 
         "Description": "로그 파일 로테이션 크기 (MB)", "Category": "logging"},
        {"Key": "MAX_LOG_FILES", "Value": 5, "Type": "int", 
         "Description": "최대 로그 파일 수", "Category": "logging"},
    ]
    
    # DataFrame으로 변환
    return {
        "Core_Constants": pd.DataFrame(core_constants_data),
        "Magic_Numbers": pd.DataFrame(magic_numbers_data),
        "Test_Constants": pd.DataFrame(test_constants_data),
        "UI_Constants": pd.DataFrame(ui_constants_data),
        "Performance_Constants": pd.DataFrame(performance_constants_data),
    }


def add_constants_sheets_to_excel(excel_path: Path) -> None:
    """엑셀 파일에 상수 시트들을 추가합니다."""
    
    print(f"📊 엑셀 파일에 상수 시트 추가 중: {excel_path}")
    
    # 상수 시트 데이터 생성
    constants_sheets = create_constants_sheets_data()
    
    try:
        # 기존 엑셀 파일 로드
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            
            for sheet_name, df in constants_sheets.items():
                print(f"  ✅ {sheet_name} 시트 추가 중... ({len(df)}개 상수)")
                
                # 시트에 데이터 쓰기
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 스타일링 적용
                worksheet = writer.sheets[sheet_name]
                apply_constants_sheet_styling(worksheet, df)
        
        print("🎉 모든 상수 시트가 성공적으로 추가되었습니다!")
        
        # 추가된 시트 목록 출력
        print("\n📋 추가된 상수 시트들:")
        for sheet_name, df in constants_sheets.items():
            print(f"  - {sheet_name}: {len(df)}개 상수")
        
    except Exception as e:
        print(f"❌ 엑셀 파일 업데이트 실패: {e}")
        raise


def apply_constants_sheet_styling(worksheet, df: pd.DataFrame) -> None:
    """상수 시트에 스타일링을 적용합니다."""
    
    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 행 스타일링
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 컬럼 너비 자동 조정
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 데이터 행 스타일링 (교대로 색상 적용)
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row_num in range(2, len(df) + 2):
        if row_num % 2 == 0:  # 짝수 행
            for col_num in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_num, column=col_num).fill = light_fill


def main():
    """메인 실행 함수"""
    excel_path = Path("data/game_initial_values_with_formulas.xlsx")
    
    if not excel_path.exists():
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return
    
    print("🚀 상수 관리 시스템 구축 시작!")
    print("=" * 50)
    
    try:
        add_constants_sheets_to_excel(excel_path)
        
        print("\n" + "=" * 50)
        print("✨ 상수 관리 시스템 구축 완료!")
        print("\n📖 사용법:")
        print("1. 엑셀 파일을 열어서 새로 추가된 상수 시트들을 확인하세요")
        print("2. 필요한 상수 값들을 수정하세요")
        print("3. 코드에서 GameConstants 클래스를 사용하여 상수에 접근하세요")
        print("\n💡 예시:")
        print("from core.adapters.excel_constants_provider import GameConstants, get_constants_provider")
        print("provider = get_constants_provider()")
        print("threshold = GameConstants.PROBABILITY_HIGH_THRESHOLD.get(provider)")
        
    except Exception as e:
        print(f"💥 오류 발생: {e}")
        return


if __name__ == "__main__":
    main() 